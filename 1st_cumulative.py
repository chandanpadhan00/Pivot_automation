from pathlib import Path
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook

# ---------------- CONFIG ----------------
CSV_PATH = r"C:\path\to\your\source.csv"          # <-- change
OUTPUT_XLSX = r"C:\path\to\your\output.xlsx"      # <-- change (can be same file daily if you want)
SOURCE_SHEET = "Source"
PIVOT_SHEET = "Pivot_Summary"

# Column names in CSV (change only if your csv uses different names)
COL_DRUG = "drug"
COL_REASON = "case_sub_status_reason_code"
COL_COUNT = "case_count"
# ----------------------------------------


def autosize_columns(ws, max_col=3):
    """Auto-fit column width based on cell content length."""
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


def format_sheet(ws):
    """Apply basic formatting to make it look professional."""
    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    # Align text columns left, numbers right
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in (1, 2):
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")

    autosize_columns(ws, max_col=3)


def build_excel_like_pivot(df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates an Excel-like pivot table output:
    - For each drug, show total row
    - Then list each reason and its sum
    - Finally grand total row
    """
    # Pivot detail rows
    detail = (
        df.groupby([COL_DRUG, COL_REASON], as_index=False)[COL_COUNT]
        .sum()
        .sort_values([COL_DRUG, COL_COUNT], ascending=[True, False])
    )

    # Drug totals
    totals = (
        df.groupby(COL_DRUG, as_index=False)[COL_COUNT]
        .sum()
        .sort_values(COL_COUNT, ascending=False)
    )

    rows = []
    grand_total = df[COL_COUNT].sum()

    # Build output rows
    for _, t in totals.iterrows():
        drug = t[COL_DRUG]
        drug_total = int(t[COL_COUNT])

        # Drug header/total row
        rows.append([drug, "", drug_total])

        # Reasons under drug
        sub = detail[detail[COL_DRUG] == drug]
        for _, r in sub.iterrows():
            rows.append(["", "  " + str(r[COL_REASON]), int(r[COL_COUNT])])

        # Blank spacer row
        rows.append(["", "", ""])

    # Grand total row
    rows.append(["Grand Total", "", int(grand_total)])

    out = pd.DataFrame(rows, columns=[COL_DRUG, COL_REASON, COL_COUNT])
    return out


def main():
    csv_path = Path(CSV_PATH)
    out_path = Path(OUTPUT_XLSX)

    if not csv_path.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}")

    # Read CSV
    df = pd.read_csv(csv_path)

    # Validate required columns
    required = {COL_DRUG, COL_REASON, COL_COUNT}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns in CSV: {missing}. Found: {list(df.columns)}")

    # Cleanup
    df[COL_DRUG] = df[COL_DRUG].astype(str).str.strip()
    df[COL_REASON] = df[COL_REASON].astype(str).str.strip()
    df[COL_COUNT] = pd.to_numeric(df[COL_COUNT], errors="coerce").fillna(0)

    # Build pivot output (Excel-like)
    pivot_out = build_excel_like_pivot(df)

    # Write to Excel (fix for if_sheet_exists)
    mode = "a" if out_path.exists() else "w"
    writer_kwargs = dict(engine="openpyxl", mode=mode)
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "replace"

    with pd.ExcelWriter(out_path, **writer_kwargs) as writer:
        df.to_excel(writer, sheet_name=SOURCE_SHEET, index=False)
        pivot_out.to_excel(writer, sheet_name=PIVOT_SHEET, index=False)

    # Apply formatting via openpyxl
    wb = load_workbook(out_path)
    ws_pivot = wb[PIVOT_SHEET]
    ws_source = wb[SOURCE_SHEET]

    format_sheet(ws_source)
    format_sheet(ws_pivot)

    # Make "Grand Total" bold
    for row in ws_pivot.iter_rows():
        if row[0].value == "Grand Total":
            for c in row[:3]:
                c.font = Font(bold=True)

    wb.save(out_path)
    print(f"✅ Done. Pivot created in: {out_path} (Sheet: {PIVOT_SHEET})")


if __name__ == "__main__":
    main()

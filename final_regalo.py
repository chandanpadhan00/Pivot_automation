from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ===================== CONFIG =====================
CSV1_PATH = r"C:\path\to\source_1.csv"          # report 1 source
CSV2_PATH = r"C:\path\to\source_2.csv"          # report 2 source
OUTPUT_XLSX = r"C:\path\to\final_output.xlsx"   # single combined output

# ---- Report 1 sheet names ----
R1_SOURCE_SHEET = "Cumulative Regalo Pending"
R1_PIVOT_SHEET  = "Summary"

# ---- Report 1 columns ----
R1_COL_DRUG   = "drug"
R1_COL_REASON = "case_sub_status_reason_code"
R1_COL_COUNT  = "case_count"

# ---- Report 2 sheet names ----
R2_SHEET1 = "All Pending Cases by Case ID"
R2_SHEET2 = "Aging by Status & Drug"

# ---- Report 2 columns ----
R2_COL_DRUG       = "drug"
R2_COL_STATUS     = "case_sub_status"
R2_COL_REASON     = "case_sub_status_reason_code"
R2_COL_FILE_RCPT  = "file_receipt_date_time"
R2_COL_ELIG_START = "eligibility_start_date"

R2_DAYS_COL = "File Receipt Date Until Today"
R2_BUCKETS = [
    ("0-15", 0, 15),
    ("15-30", 15, 30),
    ("30-45", 30, 45),
    ("45-60", 45, 60),
    ("60-75", 60, 75),
    ("75-90", 75, 90),
    ("90+", 90, None)  # >90
]
# ==================================================


# ------------------ Formatting helpers ------------------
def autosize_columns(ws, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def format_sheet_basic(ws, freeze_cell="A2"):
    # header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # align: numbers right, text left
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                continue
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = freeze_cell
    autosize_columns(ws)


# ------------------ Report 1 logic ------------------
def r1_clean_reason(text):
    if pd.isna(text):
        return "Unknown"
    text = str(text).strip()
    text = " ".join(text.split())
    text = text.lower()
    return text.title()


def build_r1_excel_like_pivot(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df[R1_COL_REASON] = df[R1_COL_REASON].apply(r1_clean_reason)

    detail = (
        df.groupby([R1_COL_DRUG, R1_COL_REASON], as_index=False)[R1_COL_COUNT]
        .sum()
        .sort_values([R1_COL_DRUG, R1_COL_REASON], ascending=[True, True])
    )

    totals = (
        df.groupby(R1_COL_DRUG, as_index=False)[R1_COL_COUNT]
        .sum()
        .sort_values(R1_COL_DRUG, ascending=True)
    )

    rows = []
    grand_total = df[R1_COL_COUNT].sum()

    for _, t in totals.iterrows():
        drug = t[R1_COL_DRUG]
        drug_total = int(t[R1_COL_COUNT])
        rows.append([drug, "", drug_total])

        sub = detail[detail[R1_COL_DRUG] == drug]
        for _, r in sub.iterrows():
            rows.append(["", "  " + str(r[R1_COL_REASON]), int(r[R1_COL_COUNT])])

        rows.append(["", "", ""])

    rows.append(["Grand Total", "", int(grand_total)])
    return pd.DataFrame(rows, columns=[R1_COL_DRUG, R1_COL_REASON, R1_COL_COUNT])


def run_report_1(csv_path: Path):
    df = pd.read_csv(csv_path)

    required = {R1_COL_DRUG, R1_COL_REASON, R1_COL_COUNT}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"[Report1] Missing columns: {missing}. Found: {list(df.columns)}")

    df[R1_COL_DRUG] = df[R1_COL_DRUG].astype(str).str.strip()
    df[R1_COL_REASON] = df[R1_COL_REASON].astype(str).str.strip()
    df[R1_COL_COUNT] = pd.to_numeric(df[R1_COL_COUNT], errors="coerce").fillna(0)

    pivot_out = build_r1_excel_like_pivot(df)
    return df, pivot_out


# ------------------ Report 2 logic ------------------
def r2_clean_text(x, unknown="Unknown"):
    if pd.isna(x):
        return unknown
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return unknown
    s = " ".join(s.split())
    s = s.lower()
    return s.title()


def build_r2_sheet1(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df[R2_COL_FILE_RCPT] = pd.to_datetime(df.get(R2_COL_FILE_RCPT), errors="coerce")
    df[R2_COL_ELIG_START] = pd.to_datetime(df.get(R2_COL_ELIG_START), errors="coerce")

    chosen_date = df[R2_COL_FILE_RCPT].where(df[R2_COL_FILE_RCPT].notna(), df[R2_COL_ELIG_START])
    today = pd.Timestamp.today().normalize()

    days = (today - chosen_date).dt.days
    days = days.where(days.notna(), other=pd.NA).clip(lower=0)

    df[R2_DAYS_COL] = days

    # bucket columns as 1 or blank "" (Excel-like)
    for name, lo, hi in R2_BUCKETS:
        if hi is None:
            mask = df[R2_DAYS_COL].notna() & (df[R2_DAYS_COL] > lo)
        else:
            mask = df[R2_DAYS_COL].notna() & (df[R2_DAYS_COL] >= lo) & (df[R2_DAYS_COL] < hi)

        df[name] = ""
        df.loc[mask, name] = 1

    return df


def build_r2_sheet2_pivot(df_sheet1: pd.DataFrame) -> pd.DataFrame:
    df = df_sheet1.copy()

    df[R2_COL_DRUG] = df[R2_COL_DRUG].astype(str).str.strip()
    df[R2_COL_STATUS] = df[R2_COL_STATUS].apply(r2_clean_text)
    df[R2_COL_REASON] = df[R2_COL_REASON].apply(r2_clean_text)

    # numeric versions of buckets for sums
    for name, _, _ in R2_BUCKETS:
        df[name] = df[name].apply(lambda v: 1 if v == 1 else 0)

    pivot = pd.pivot_table(
        df,
        index=[R2_COL_DRUG, R2_COL_STATUS, R2_COL_REASON],
        values=[b[0] for b in R2_BUCKETS],
        aggfunc="sum",
        fill_value=0
    ).sort_index(level=[0, 1, 2], ascending=[True, True, True]).reset_index()

    return pivot


def run_report_2(csv_path: Path):
    df = pd.read_csv(csv_path)

    required = {R2_COL_DRUG, R2_COL_STATUS, R2_COL_REASON, R2_COL_FILE_RCPT, R2_COL_ELIG_START}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"[Report2] Missing columns: {missing}. Found: {list(df.columns)}")

    sheet1 = build_r2_sheet1(df)
    sheet2 = build_r2_sheet2_pivot(sheet1)
    return sheet1, sheet2


# ------------------ Main writer ------------------
def write_all_sheets(out_path: Path, sheets: dict[str, pd.DataFrame]):
    mode = "a" if out_path.exists() else "w"
    writer_kwargs = dict(engine="openpyxl", mode=mode)
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "replace"

    with pd.ExcelWriter(out_path, **writer_kwargs) as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # apply formatting
    wb = load_workbook(out_path)
    for sheet_name in sheets.keys():
        ws = wb[sheet_name]
        format_sheet_basic(ws, freeze_cell="A2")

        # bold "Grand Total" row for Report1 Summary if present
        if sheet_name == R1_PIVOT_SHEET:
            for row in ws.iter_rows():
                if row[0].value == "Grand Total":
                    for c in row[:3]:
                        c.font = Font(bold=True)

    wb.save(out_path)


def main():
    csv1 = Path(CSV1_PATH)
    csv2 = Path(CSV2_PATH)
    out = Path(OUTPUT_XLSX)

    if not csv1.exists():
        raise FileNotFoundError(f"CSV1 not found: {csv1}")
    if not csv2.exists():
        raise FileNotFoundError(f"CSV2 not found: {csv2}")

    # Build all dataframes
    r1_source, r1_summary = run_report_1(csv1)
    r2_all_cases, r2_aging_pivot = run_report_2(csv2)

    # Write 4 sheets into one workbook
    sheets = {
        R1_SOURCE_SHEET: r1_source,
        R1_PIVOT_SHEET:  r1_summary,
        R2_SHEET1:       r2_all_cases,
        R2_SHEET2:       r2_aging_pivot
    }

    write_all_sheets(out, sheets)
    print(f"✅ Done. 4 sheets written to: {out}")


if __name__ == "__main__":
    main()

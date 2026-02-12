from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ---------------- CONFIG ----------------
CSV_PATH = r"C:\path\to\source.csv"              # <-- change
OUTPUT_XLSX = r"C:\path\to\output.xlsx"          # <-- change

SHEET1 = "All Pending Cases by Case ID"
SHEET2 = "Aging by Status & Drug"

# CSV column names (change only if your csv uses different names)
COL_DRUG = "drug"
COL_STATUS = "case_sub_status"
COL_REASON = "case_sub_status_reason_code"
COL_FILE_RCPT = "file_receipt_date_time"         # date
COL_ELIG_START = "eligibility_start_date"        # date

# ----------------------------------------


BUCKETS = [
    ("0-15", 0, 15),
    ("15-30", 15, 30),
    ("30-45", 30, 45),
    ("45-60", 45, 60),
    ("60-75", 60, 75),
    ("75-90", 75, 90),
    ("90+", 90, None)   # 90+ means >90
]

DAYS_COL = "File Receipt Date Until Today"


def clean_text(x, unknown="Unknown"):
    """Normalize text so Excel-like pivot grouping happens (case-insensitive + trims)."""
    if pd.isna(x):
        return unknown
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return unknown
    s = " ".join(s.split())      # remove extra internal spaces
    s = s.lower()
    return s.title()             # nice report display


def autosize_columns(ws, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def format_sheet_basic(ws, freeze_cell="A2"):
    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Align columns
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                continue
            # numbers right, text left
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = freeze_cell
    autosize_columns(ws)


def build_sheet1(df: pd.DataFrame) -> pd.DataFrame:
    """Create Sheet1 with Days + Bucket columns as 1 or blank (Excel-like)."""
    # Parse date columns (safe)
    df[COL_FILE_RCPT] = pd.to_datetime(df.get(COL_FILE_RCPT), errors="coerce")
    df[COL_ELIG_START] = pd.to_datetime(df.get(COL_ELIG_START), errors="coerce")

    # Choose date: if file_receipt_date_time blank -> eligibility_start_date else file_receipt_date_time
    chosen_date = df[COL_FILE_RCPT].where(df[COL_FILE_RCPT].notna(), df[COL_ELIG_START])

    # Days until today
    today = pd.Timestamp.today().normalize()
    days = (today - chosen_date).dt.days

    # If both dates missing, keep blank
    # If negative, clip to 0 (optional but usually correct for "until today" aging)
    days = days.where(days.notna(), other=pd.NA)
    days = days.clip(lower=0)

    df[DAYS_COL] = days

    # Create bucket columns as 1 or "" (blank)
    for name, lo, hi in BUCKETS:
        if hi is None:
            mask = df[DAYS_COL].notna() & (df[DAYS_COL] > lo)
        else:
            mask = df[DAYS_COL].notna() & (df[DAYS_COL] >= lo) & (df[DAYS_COL] < hi)

        df[name] = ""
        df.loc[mask, name] = 1

    return df


def build_sheet2_pivot(df_sheet1: pd.DataFrame) -> pd.DataFrame:
    """Create Sheet2 pivot (Aging by Status & Drug)."""

    # Normalize row fields to merge case-mismatch like Excel pivot
    df = df_sheet1.copy()
    df[COL_DRUG] = df[COL_DRUG].astype(str).str.strip()
    df[COL_STATUS] = df[COL_STATUS].apply(clean_text)
    df[COL_REASON] = df[COL_REASON].apply(clean_text)

    # For pivot sums, need numeric 1/0 versions of bucket columns
    for name, _, _ in BUCKETS:
        df[name] = df[name].apply(lambda v: 1 if v == 1 else 0)

    # Pivot: rows = drug, status, reason; values = sum of buckets
    pivot = pd.pivot_table(
        df,
        index=[COL_DRUG, COL_STATUS, COL_REASON],
        values=[b[0] for b in BUCKETS],
        aggfunc="sum",
        fill_value=0
    )

    # Sort by drug/status/reason ascending (like you asked earlier)
    pivot = pivot.sort_index(level=[0, 1, 2], ascending=[True, True, True])

    # Flatten for Excel
    pivot = pivot.reset_index()

    return pivot


def write_excel(df1: pd.DataFrame, df2: pd.DataFrame, out_path: Path):
    mode = "a" if out_path.exists() else "w"
    writer_kwargs = dict(engine="openpyxl", mode=mode)
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "replace"

    with pd.ExcelWriter(out_path, **writer_kwargs) as writer:
        df1.to_excel(writer, sheet_name=SHEET1, index=False)
        df2.to_excel(writer, sheet_name=SHEET2, index=False)

    # Formatting
    wb = load_workbook(out_path)
    ws1 = wb[SHEET1]
    ws2 = wb[SHEET2]

    format_sheet_basic(ws1, freeze_cell="A2")
    format_sheet_basic(ws2, freeze_cell="A2")

    # Make the bucket headers stand out a bit (optional)
    # (Keep simple & clean: just bold already done)

    wb.save(out_path)


def main():
    csv_path = Path(CSV_PATH)
    out_path = Path(OUTPUT_XLSX)

    if not csv_path.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}")

    df = pd.read_csv(csv_path)

    # Validate required columns
    required = {COL_DRUG, COL_STATUS, COL_REASON, COL_FILE_RCPT, COL_ELIG_START}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns in CSV: {missing}. Found: {list(df.columns)}")

    # Build sheets
    sheet1_df = build_sheet1(df)
    sheet2_df = build_sheet2_pivot(sheet1_df)

    # Write output
    write_excel(sheet1_df, sheet2_df, out_path)

    print(f"✅ Done. Created/updated:\n- {SHEET1}\n- {SHEET2}\nFile: {out_path}")


if __name__ == "__main__":
    main()
